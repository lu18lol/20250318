import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
from ..models import Question, Option, Category, Tag

class ExcelProcessor:
    TEMPLATE_HEADERS = [
        '题目类型', '题目内容', '难度', '分类', '标签', 
        '选项A', '选项B', '选项C', '选项D', 
        '正确答案', '解析'
    ]
    
    QUESTION_TYPES = {
        '单选题': 'single',
        '多选题': 'multiple',
        '判断题': 'truefalse',
        '问答题': 'essay'
    }
    
    DIFFICULTY_LEVELS = {
        '简单': 'easy',
        '中等': 'medium',
        '困难': 'hard'
    }
    
    @classmethod
    def create_template(cls):
        """创建Excel模板"""
        wb = Workbook()
        ws = wb.active
        ws.title = "题目导入模板"
        
        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # 写入表头
        for col, header in enumerate(cls.TEMPLATE_HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # 设置列宽
        ws.column_dimensions['A'].width = 15  # 题目类型
        ws.column_dimensions['B'].width = 50  # 题目内容
        ws.column_dimensions['C'].width = 10  # 难度
        ws.column_dimensions['D'].width = 20  # 分类
        ws.column_dimensions['E'].width = 20  # 标签
        for col in ['F', 'G', 'H', 'I']:  # 选项
            ws.column_dimensions[col].width = 30
        ws.column_dimensions['J'].width = 20  # 正确答案
        ws.column_dimensions['K'].width = 50  # 解析
        
        # 添加说明行
        ws.cell(row=2, column=1, value="可选值：单选题、多选题、判断题、问答题")
        ws.cell(row=2, column=3, value="可选值：简单、中等、困难")
        ws.cell(row=2, column=4, value="多个分类用逗号分隔")
        ws.cell(row=2, column=5, value="多个标签用逗号分隔")
        ws.cell(row=2, column=10, value="多选题答案用逗号分隔，判断题填是/否")
        
        # 添加示例数据
        example_data = [
            ['单选题', '以下哪个是Python的内置数据类型？', '简单', 'Python基础,数据类型', '基础概念',
             'int', 'char', 'String', 'float', 'A,D', '整数(int)和浮点数(float)是Python的内置数据类型'],
            ['多选题', '以下哪些是HTTP请求方法？', '中等', 'Web开发,HTTP协议', 'Web基础,协议',
             'GET', 'POST', 'DELETE', 'PATCH', 'A,B,C,D', 'HTTP定义了多种请求方法用于不同的操作'],
            ['判断题', 'Python中的列表是可变的数据类型', '简单', 'Python基础', '数据结构',
             '', '', '', '', '是', '列表(list)是可变序列，可以修改其内容'],
        ]
        
        for i, row_data in enumerate(example_data, 3):
            for j, value in enumerate(row_data, 1):
                ws.cell(row=i, column=j, value=value)
        
        return wb
    
    @classmethod
    def export_template(cls):
        """导出Excel模板"""
        wb = cls.create_template()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=question_template.xlsx'
        wb.save(response)
        return response
    
    @classmethod
    def import_questions(cls, file, user):
        """从Excel文件导入题目"""
        df = pd.read_excel(file, sheet_name=0)
        results = {
            'success': 0,
            'failed': 0,
            'errors': []
        }
        
        for index, row in df.iterrows():
            try:
                # 跳过空行和说明行
                if pd.isna(row['题目类型']) or row['题目类型'] in ['可选值：单选题、多选题、判断题、问答题']:
                    continue
                
                # 处理题目类型
                q_type = cls.QUESTION_TYPES.get(row['题目类型'])
                if not q_type:
                    raise ValueError(f"无效的题目类型: {row['题目类型']}")
                
                # 处理难度
                difficulty = cls.DIFFICULTY_LEVELS.get(row['难度'])
                if not difficulty:
                    raise ValueError(f"无效的难度级别: {row['难度']}")
                
                # 创建题目
                question = Question.objects.create(
                    type=q_type,
                    content=row['题目内容'],
                    difficulty=difficulty,
                    explanation=row['解析'] if not pd.isna(row['解析']) else '',
                    created_by=user
                )
                
                # 处理分类
                if not pd.isna(row['分类']):
                    categories = [c.strip() for c in str(row['分类']).split(',')]
                    for cat_name in categories:
                        category, _ = Category.objects.get_or_create(name=cat_name)
                        question.categories.add(category)
                
                # 处理标签
                if not pd.isna(row['标签']):
                    tags = [t.strip() for t in str(row['标签']).split(',')]
                    for tag_name in tags:
                        tag, _ = Tag.objects.get_or_create(name=tag_name)
                        question.tags.add(tag)
                
                # 处理选项和答案
                if q_type in ['single', 'multiple']:
                    correct_answers = str(row['正确答案']).strip().split(',')
                    for i, option_col in enumerate(['选项A', '选项B', '选项C', '选项D']):
                        if not pd.isna(row[option_col]):
                            is_correct = chr(65 + i) in correct_answers
                            Option.objects.create(
                                question=question,
                                content=row[option_col],
                                is_correct=is_correct
                            )
                elif q_type == 'truefalse':
                    Option.objects.create(
                        question=question,
                        content='正确',
                        is_correct=row['正确答案'] == '是'
                    )
                
                results['success'] += 1
            except Exception as e:
                results['failed'] += 1
                results['errors'].append(f"第{index+1}行: {str(e)}")
        
        return results 